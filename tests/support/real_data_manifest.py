"""Canonical, checkout-relative inventory for ignored real-data fixtures."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
MANIFEST_PATH = PROJECT_ROOT / "tests" / "real_data" / "reference_data_manifest.json"


class ReferenceDataManifestError(ValueError):
    """Raised when the tracked reference-data contract is invalid."""


@dataclass(frozen=True)
class ReferenceDataFile:
    id: str
    filename: str
    sha256: str
    bytes: int
    rows: int
    sheets: tuple[str, ...]
    required_headers: frozenset[str]
    uses: tuple[str, ...]


@dataclass(frozen=True)
class ReferenceDataManifest:
    root: Path
    files: tuple[ReferenceDataFile, ...]

    @property
    def by_id(self) -> dict[str, ReferenceDataFile]:
        return {item.id: item for item in self.files}

    @property
    def by_filename(self) -> dict[str, ReferenceDataFile]:
        return {item.filename: item for item in self.files}

    def path(self, filename_or_id: str) -> Path:
        item = self.by_id.get(filename_or_id) or self.by_filename.get(filename_or_id)
        if item is None:
            raise KeyError(f"unknown reference-data file: {filename_or_id}")
        return self.root / item.filename


def _required_string(value: Any, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ReferenceDataManifestError(f"{field} must be a non-empty string")
    return value


def _string_list(value: Any, field: str) -> tuple[str, ...]:
    if not isinstance(value, list) or not value:
        raise ReferenceDataManifestError(f"{field} must be a non-empty list")
    result = tuple(_required_string(item, field) for item in value)
    if len(result) != len(set(result)):
        raise ReferenceDataManifestError(f"{field} contains duplicates")
    return result


def load_reference_data_manifest(path: Path = MANIFEST_PATH) -> ReferenceDataManifest:
    """Load and validate the one tracked reference-data inventory."""

    path = Path(path).resolve()
    raw = json.loads(path.read_text(encoding="utf-8"))
    if raw.get("schema_version") != "reference_data_manifest.v1":
        raise ReferenceDataManifestError("unsupported reference-data manifest schema")

    relative_root = Path(_required_string(raw.get("root"), "root"))
    if relative_root.is_absolute() or ".." in relative_root.parts:
        raise ReferenceDataManifestError("root must stay within the checkout")
    root = (PROJECT_ROOT / relative_root).resolve()
    if PROJECT_ROOT not in root.parents:
        raise ReferenceDataManifestError("root escapes the checkout")

    entries = raw.get("files")
    if not isinstance(entries, list) or not entries:
        raise ReferenceDataManifestError("files must be a non-empty list")

    files: list[ReferenceDataFile] = []
    for index, entry in enumerate(entries):
        if not isinstance(entry, dict):
            raise ReferenceDataManifestError(f"files[{index}] must be an object")
        file_id = _required_string(entry.get("id"), f"files[{index}].id")
        filename = _required_string(entry.get("filename"), f"files[{index}].filename")
        if Path(filename).name != filename:
            raise ReferenceDataManifestError(f"files[{index}].filename must be a basename")
        sha256 = _required_string(entry.get("sha256"), f"files[{index}].sha256")
        if len(sha256) != 64 or any(char not in "0123456789abcdef" for char in sha256):
            raise ReferenceDataManifestError(f"files[{index}].sha256 is invalid")
        expected_bytes = entry.get("bytes")
        rows = entry.get("rows")
        if not isinstance(expected_bytes, int) or expected_bytes <= 0:
            raise ReferenceDataManifestError(f"files[{index}].bytes must be positive")
        if not isinstance(rows, int) or rows <= 0:
            raise ReferenceDataManifestError(f"files[{index}].rows must be positive")
        files.append(
            ReferenceDataFile(
                id=file_id,
                filename=filename,
                sha256=sha256,
                bytes=expected_bytes,
                rows=rows,
                sheets=_string_list(entry.get("sheets"), f"files[{index}].sheets"),
                required_headers=frozenset(
                    _string_list(entry.get("required_headers"), f"files[{index}].required_headers")
                ),
                uses=_string_list(entry.get("uses"), f"files[{index}].uses"),
            )
        )

    ids = [item.id for item in files]
    filenames = [item.filename for item in files]
    if len(ids) != len(set(ids)):
        raise ReferenceDataManifestError("file ids must be unique")
    if len(filenames) != len(set(filenames)):
        raise ReferenceDataManifestError("filenames must be unique")
    return ReferenceDataManifest(root=root, files=tuple(files))


REFERENCE_DATA = load_reference_data_manifest()
REFERENCE_DATA_DIR = REFERENCE_DATA.root
REFERENCE_DATA_AVAILABLE = REFERENCE_DATA_DIR.is_dir()


def reference_data_path(filename_or_id: str) -> Path:
    """Return a canonical workbook path without fallback aliases."""

    return REFERENCE_DATA.path(filename_or_id)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_reference_data(*, verify_hashes: bool = True) -> list[str]:
    """Return precise contract errors for the current ignored fixture set."""

    errors: list[str] = []
    if not REFERENCE_DATA_DIR.is_dir():
        return [f"reference-data directory is missing: {REFERENCE_DATA_DIR}"]

    actual_names = {path.name for path in REFERENCE_DATA_DIR.glob("*.xlsx")}
    expected_names = set(REFERENCE_DATA.by_filename)
    missing = sorted(expected_names - actual_names)
    unexpected = sorted(actual_names - expected_names)
    if missing:
        errors.append(f"missing reference-data files: {missing}")
    if unexpected:
        errors.append(f"unexpected reference-data files: {unexpected}")

    for item in REFERENCE_DATA.files:
        path = REFERENCE_DATA_DIR / item.filename
        if not path.is_file():
            continue
        actual_bytes = path.stat().st_size
        if actual_bytes != item.bytes:
            errors.append(f"{item.filename}: bytes {actual_bytes} != {item.bytes}")
        if verify_hashes:
            actual_hash = _sha256(path)
            if actual_hash != item.sha256:
                errors.append(f"{item.filename}: sha256 {actual_hash} != {item.sha256}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if tuple(workbook.sheetnames) != item.sheets:
                errors.append(
                    f"{item.filename}: sheets {tuple(workbook.sheetnames)} != {item.sheets}"
                )
            sheet = workbook[workbook.sheetnames[0]]
            headers = {
                str(value).strip()
                for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if value is not None and str(value).strip()
            }
            absent_headers = sorted(item.required_headers - headers)
            if absent_headers:
                errors.append(f"{item.filename}: missing headers {absent_headers}")
            actual_rows = max(sheet.max_row - 1, 0)
            if actual_rows != item.rows:
                errors.append(f"{item.filename}: rows {actual_rows} != {item.rows}")
        finally:
            workbook.close()
    return errors
