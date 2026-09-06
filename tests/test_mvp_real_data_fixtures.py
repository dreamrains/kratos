from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.support.real_data_manifest import (
    REFERENCE_DATA_AVAILABLE,
    REFERENCE_DATA_DIR,
    reference_data_path,
)

TEST_DOC_DIR = REFERENCE_DATA_DIR


EXPECTED_FILES = {
    "游戏B留存.xlsx": {"日期", "日活跃", "日新增", "1天后", "7天后"},
    "游戏A内购数据.xlsx": {"日期", "活跃用户", "付费人数", "内购收入", "付费率"},
    "省钱卡订单.xlsx": {"user_id", "商品名称", "售价", "支付时间"},
}


def _headers(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            values = {str(value).strip() for value in row if value is not None and str(value).strip()}
            if values:
                return values
    finally:
        wb.close()
    return set()


@pytest.mark.skipif(not REFERENCE_DATA_AVAILABLE, reason="canonical reference data is not installed")
def test_fast_real_data_fixtures_are_available_and_readable():
    for filename, required_headers in EXPECTED_FILES.items():
        path = TEST_DOC_DIR / filename
        assert path.exists(), f"{filename} is missing"
        headers = _headers(path)
        assert required_headers <= headers


@pytest.mark.skipif(not REFERENCE_DATA_AVAILABLE, reason="canonical reference data is not installed")
def test_large_real_data_fixture_is_present_but_not_loaded_by_fast_tests():
    path = reference_data_path("savings_card_user_payments")

    assert path.exists()
    assert path.stat().st_size > 500_000
